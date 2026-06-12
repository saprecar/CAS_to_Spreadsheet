import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

def generate_excel(parsed_data):
    metadata = parsed_data.get("metadata", {})
    holdings = parsed_data.get("holdings", [])
    
    wb = Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary & Metadata"
    
    # Styling
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    
    # Add Metadata
    ws_summary.append(["Investor Details", ""])
    ws_summary["A1"].font = header_font
    ws_summary["A1"].fill = header_fill
    ws_summary["B1"].fill = header_fill
    
    ws_summary.append(["Name", metadata.get("investor_name", "N/A")])
    ws_summary.append(["Email", metadata.get("email", "N/A")])
    ws_summary.append(["Mobile", metadata.get("mobile", "N/A")])
    ws_summary.append(["Address", metadata.get("address", "N/A")])
    ws_summary.append(["Statement Date", metadata.get("statement_date", "N/A")])
    ws_summary.append([])
    
    # Add Portfolio Metrics
    ws_summary.append(["Portfolio Metrics", ""])
    r = ws_summary.max_row
    ws_summary.cell(row=r, column=1).font = header_font
    ws_summary.cell(row=r, column=1).fill = header_fill
    ws_summary.cell(row=r, column=2).fill = header_fill
    
    total_market = sum(h["market_value"] for h in holdings)
    total_cost = sum(h["cost_value"] for h in holdings)
    total_gain = total_market - total_cost
    gain_percent = (total_gain / total_cost * 100) if total_cost > 0 else 0
    
    ws_summary.append(["Total Market Value", total_market])
    ws_summary.append(["Total Cost Value", total_cost])
    ws_summary.append(["Absolute Gain", total_gain])
    ws_summary.append(["Gain %", gain_percent])
    
    for row in range(r+1, r+4):
        ws_summary.cell(row=row, column=2).number_format = '₹#,##0.00'
    ws_summary.cell(row=r+4, column=2).number_format = '0.00%'
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40
    
    # 2. Holdings Sheet
    ws_holdings = wb.create_sheet(title="Holdings Valuation")
    
    if holdings:
        df = pd.DataFrame(holdings)
        # Drop transactions column if it exists to avoid pandas expanding it
        if 'transactions' in df.columns:
            df = df.drop(columns=['transactions'])
            
        cols = ['folio', 'registrar', 'scheme_code', 'scheme_name', 'isin', 'demat_status', 
                'units', 'nav_date', 'nav', 'cost_value', 'market_value', 'absolute_gain', 'gain_percent']
        df = df[[c for c in cols if c in df.columns]]
        
        df.columns = ['Folio No', 'Registrar', 'Scheme Code', 'Scheme Name', 'ISIN', 'Demat Status',
                      'Units', 'NAV Date', 'NAV', 'Cost Value', 'Market Value', 'Absolute Gain', 'Gain %']
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_holdings.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx in [7]: # Units
                        cell.number_format = '#,##0.000'
                    elif c_idx in [9]: # NAV
                        cell.number_format = '#,##0.0000'
                    elif c_idx in [10, 11, 12]: # Cost, Market, Gain
                        cell.number_format = '₹#,##0.00'
                    elif c_idx == 13: # Gain %
                        if cell.value is not None:
                            cell.value = cell.value / 100
                            cell.number_format = '0.00%'
                        
                    if c_idx in [12, 13]:
                        if isinstance(cell.value, (int, float)):
                            if cell.value > 0:
                                cell.font = Font(color="008000")
                            elif cell.value < 0:
                                cell.font = Font(color="FF0000")
        
        max_r = ws_holdings.max_row
        if max_r > 1:
            ws_holdings.cell(row=max_r+1, column=1, value="TOTAL").font = bold_font
            ws_holdings.cell(row=max_r+1, column=10, value=f"=SUM(J2:J{max_r})").font = bold_font
            ws_holdings.cell(row=max_r+1, column=11, value=f"=SUM(K2:K{max_r})").font = bold_font
            ws_holdings.cell(row=max_r+1, column=12, value=f"=SUM(L2:L{max_r})").font = bold_font
            
            ws_holdings.cell(row=max_r+1, column=10).number_format = '₹#,##0.00'
            ws_holdings.cell(row=max_r+1, column=11).number_format = '₹#,##0.00'
            ws_holdings.cell(row=max_r+1, column=12).number_format = '₹#,##0.00'
            
            thick_border = Border(top=Side(style='thin'), bottom=Side(style='double'))
            for c in range(1, 14):
                ws_holdings.cell(row=max_r+1, column=c).border = thick_border
            
        for col in ws_holdings.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            ws_holdings.column_dimensions[column].width = adjusted_width

    # 3. Transactions Sheet (if available)
    has_transactions = any(len(h.get("transactions", [])) > 0 for h in holdings)
    if has_transactions:
        ws_tx = wb.create_sheet(title="Transactions")
        all_txs = []
        for h in holdings:
            for tx in h.get("transactions", []):
                all_txs.append({
                    "Mutual Fund": h.get("mutual_fund", ""),
                    "Folio No": h.get("folio", ""),
                    "Scheme Name": h.get("scheme_name", ""),
                    "Date": tx.get("date", ""),
                    "Description": tx.get("desc", ""),
                    "Amount (INR)": tx.get("amount"),
                    "Price / NAV": tx.get("price"),
                    "Units": tx.get("units"),
                    "Unit Balance": tx.get("balance")
                })
                
        df_tx = pd.DataFrame(all_txs)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_tx, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_tx.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx == 6: # Amount
                        if isinstance(value, (int, float)):
                            cell.number_format = '₹#,##0.00'
                            if value < 0:
                                cell.font = Font(color="FF0000")
                    elif c_idx == 7: # Price
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.0000'
                    elif c_idx in [8, 9]: # Units, Balance
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.000'
                            if value < 0:
                                cell.font = Font(color="FF0000")
                                
        for col in ws_tx.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 60:
                adjusted_width = 60
            ws_tx.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
